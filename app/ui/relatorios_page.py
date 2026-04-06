"""Página de Relatórios com análises inteligentes e drill-down por nota."""
from __future__ import annotations

from datetime import datetime, date, time
from pathlib import Path
from collections import defaultdict

from PySide6.QtCore import QDate, Qt
from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
    QTableWidget, QTableWidgetItem, QGroupBox, QMessageBox, QTabWidget, QFileDialog,
    QHeaderView, QProgressBar, QLineEdit, QDialog, QAbstractItemView, QDateEdit,
    QDialogButtonBox
)

from app.core import app_signals
from app.db import get_db_session, EmpresaRepository, DocumentoRepository, CredencialRepository
from app.utils.logger import log
from app.utils.ui_state import get_settings, load_qdate, save_qdate
from app.utils.document_viewer import open_document_file

# Variável global para armazenar a empresa selecionada
_selected_empresa_id = None

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class RelatoriosPage(QWidget):
    """Página de relatórios com análises inteligentes e pesquisa por aba."""

    def __init__(self):
        super().__init__()
        self.session = get_db_session()
        self.empresa_repo = EmpresaRepository(self.session)
        self.doc_repo = DocumentoRepository(self.session)
        self.cred_repo = CredencialRepository(self.session)
        self.settings = get_settings()
        self._restoring_state = False
        self.dados_notas = []
        self.empresa_selecionada_id = None
        self.empresa_cnpj = ""
        self.empresa_nome = ""
        self._resultados_cliente = []
        self._resultados_fornecedor = []
        self._resultados_servico = []
        self._resultados_impostos = []
        self._resultados_tendencias = []
        self._resultados_cancelamentos = []
        self._user_generated_once = False
        self._pending_auto_refresh = False

        layout = QVBoxLayout(self)

        # Título
        title = QLabel("📊 Relatórios Inteligentes")
        title.setStyleSheet("font-size: 18px; font-weight: bold; margin-bottom: 10px;")
        layout.addWidget(title)

        subtitle = QLabel(
            "Análises profissionais automáticas de suas notas fiscais. "
            "Selecione o período para gerar os relatórios."
        )
        subtitle.setWordWrap(True)
        subtitle.setStyleSheet("color: #64748b; margin-bottom: 10px;")
        layout.addWidget(subtitle)

        # Painel de filtros
        filter_group = QGroupBox("Filtros")
        filter_layout = QVBoxLayout(filter_group)

        empresa_row = QHBoxLayout()
        empresa_label_text = QLabel("Empresa:")
        empresa_label_text.setMinimumWidth(80)
        self.empresa_label = QLabel("Carregando...")
        self.empresa_label.setStyleSheet("font-weight: bold; color: #2563eb;")
        empresa_row.addWidget(empresa_label_text)
        empresa_row.addWidget(self.empresa_label)
        empresa_row.addStretch()
        filter_layout.addLayout(empresa_row)

        data_row = QHBoxLayout()

        data_ini_label = QLabel("Data Inicial:")
        data_ini_label.setMinimumWidth(80)
        self.data_inicial = QDateEdit()
        self.data_inicial.setDate(QDate.currentDate().addMonths(-1))
        self.data_inicial.setCalendarPopup(True)
        self.data_inicial.setDisplayFormat("dd/MM/yyyy")
        self.data_inicial.setMaximumWidth(150)
        data_row.addWidget(data_ini_label)
        data_row.addWidget(self.data_inicial)

        data_fim_label = QLabel("Data Final:")
        data_fim_label.setMinimumWidth(80)
        self.data_final = QDateEdit()
        self.data_final.setDate(QDate.currentDate())
        self.data_final.setCalendarPopup(True)
        self.data_final.setDisplayFormat("dd/MM/yyyy")
        self.data_final.setMaximumWidth(150)
        data_row.addWidget(data_fim_label)
        data_row.addWidget(self.data_final)
        data_row.addStretch()

        filter_layout.addLayout(data_row)

        btn_layout = QHBoxLayout()
        self.btn_gerar = QPushButton("🔄 Gerar Relatórios")
        self.btn_gerar.setMaximumWidth(150)
        self.btn_gerar.clicked.connect(self._gerar_relatorios)
        btn_layout.addWidget(self.btn_gerar)

        self.btn_exportar = QPushButton("📥 Exportar para Excel")
        self.btn_exportar.setMaximumWidth(150)
        self.btn_exportar.clicked.connect(self._exportar_excel)
        self.btn_exportar.setEnabled(False)
        btn_layout.addWidget(self.btn_exportar)

        self.progress = QProgressBar()
        self.progress.setValue(0)
        self.progress.setVisible(False)
        self.progress.setMaximumHeight(20)
        btn_layout.addWidget(self.progress)
        btn_layout.addStretch()

        filter_layout.addLayout(btn_layout)
        layout.addWidget(filter_group)

        # Pesquisa inteligente
        search_layout = QHBoxLayout()
        lbl_search = QLabel("Pesquisa da aba atual:")
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText(
            "Digite cliente, número, chave, valor, status... a busca vale só para a aba aberta"
        )
        self.search_input.textChanged.connect(self._aplicar_pesquisa_aba_atual)
        self.btn_limpar_pesquisa = QPushButton("Limpar pesquisa")
        self.btn_limpar_pesquisa.clicked.connect(self._limpar_pesquisa)
        search_layout.addWidget(lbl_search)
        search_layout.addWidget(self.search_input, 1)
        search_layout.addWidget(self.btn_limpar_pesquisa)
        layout.addLayout(search_layout)

        # Abas de relatórios
        self.tabs = QTabWidget()

        self.tab_financeiro = self._criar_tab_financeiro()
        self.tab_clientes = self._criar_tab_clientes()
        self.tab_fornecedores = self._criar_tab_fornecedores()
        self.tab_servicos = self._criar_tab_servicos()
        self.tab_impostos = self._criar_tab_impostos()
        self.tab_tendencias = self._criar_tab_tendencias()
        self.tab_cancelamentos = self._criar_tab_cancelamentos()

        self.tabs.addTab(self.tab_financeiro, "📈 Financeiro")
        self.tabs.addTab(self.tab_clientes, "👥 Clientes")
        self.tabs.addTab(self.tab_fornecedores, "🏭 Fornecedores")
        self.tabs.addTab(self.tab_servicos, "🛠️ Serviços")
        self.tabs.addTab(self.tab_impostos, "💰 Impostos")
        self.tabs.addTab(self.tab_tendencias, "📊 Tendências")
        self.tabs.addTab(self.tab_cancelamentos, "⚠️ Cancelamentos")
        self.tabs.currentChanged.connect(lambda _idx: self._aplicar_pesquisa_aba_atual())

        layout.addWidget(self.tabs)

        # Status
        self.status_label = QLabel("Pronto para gerar relatórios")
        self.status_label.setStyleSheet("color: #4CAF50; font-weight: bold;")
        layout.addWidget(self.status_label)

        app_signals.company_selected.connect(self._on_global_empresa_changed)
        self._restore_state()

    def _configurar_tabela(self, table: QTableWidget):
        table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        table.setSelectionBehavior(QAbstractItemView.SelectRows)
        table.setSelectionMode(QAbstractItemView.SingleSelection)
        table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        table.setAlternatingRowColors(True)

    def _criar_tab_financeiro(self) -> QWidget:
        widget = QWidget()
        layout = QVBoxLayout(widget)
        self.table_financeiro = QTableWidget()
        self.table_financeiro.setColumnCount(2)
        self.table_financeiro.setHorizontalHeaderLabels(["Métrica", "Valor"])
        self._configurar_tabela(self.table_financeiro)
        layout.addWidget(self.table_financeiro)
        return widget

    def _criar_tab_clientes(self) -> QWidget:
        widget = QWidget()
        layout = QVBoxLayout(widget)
        info = QLabel("Clique em um cliente para abrir as notas dele.")
        info.setStyleSheet("color: #64748b; font-size: 11px;")
        layout.addWidget(info)
        self.table_clientes = QTableWidget()
        self.table_clientes.setColumnCount(4)
        self.table_clientes.setHorizontalHeaderLabels(["Cliente", "Valor Total", "Qtd Notas", "Valor Médio"])
        self._configurar_tabela(self.table_clientes)
        self.table_clientes.cellClicked.connect(self._abrir_detalhe_cliente)
        layout.addWidget(self.table_clientes)
        return widget

    def _criar_tab_fornecedores(self) -> QWidget:
        widget = QWidget()
        layout = QVBoxLayout(widget)
        info = QLabel("Clique em um fornecedor para abrir as notas tomadas dele.")
        info.setStyleSheet("color: #64748b; font-size: 11px;")
        layout.addWidget(info)
        self.table_fornecedores = QTableWidget()
        self.table_fornecedores.setColumnCount(4)
        self.table_fornecedores.setHorizontalHeaderLabels(["Fornecedor", "Valor Total", "Qtd Notas", "Valor Médio"])
        self._configurar_tabela(self.table_fornecedores)
        self.table_fornecedores.cellClicked.connect(self._abrir_detalhe_fornecedor)
        layout.addWidget(self.table_fornecedores)
        return widget

    def _criar_tab_servicos(self) -> QWidget:
        widget = QWidget()
        layout = QVBoxLayout(widget)
        info = QLabel("Clique em um serviço para abrir todas as notas desse serviço.")
        info.setStyleSheet("color: #64748b; font-size: 11px;")
        layout.addWidget(info)
        self.table_servicos = QTableWidget()
        self.table_servicos.setColumnCount(3)
        self.table_servicos.setHorizontalHeaderLabels(["Serviço", "Valor Total", "Qtd Notas"])
        self._configurar_tabela(self.table_servicos)
        self.table_servicos.cellClicked.connect(self._abrir_detalhe_servico)
        layout.addWidget(self.table_servicos)
        return widget

    def _criar_tab_impostos(self) -> QWidget:
        widget = QWidget()
        layout = QVBoxLayout(widget)
        info = QLabel("Clique em um imposto para abrir as notas que possuem esse imposto destacado.")
        info.setStyleSheet("color: #64748b; font-size: 11px;")
        layout.addWidget(info)
        self.table_impostos = QTableWidget()
        self.table_impostos.setColumnCount(3)
        self.table_impostos.setHorizontalHeaderLabels(["Imposto", "Valor Total", "Qtd Notas"])
        self._configurar_tabela(self.table_impostos)
        self.table_impostos.cellClicked.connect(self._abrir_detalhe_imposto)
        layout.addWidget(self.table_impostos)
        return widget

    def _criar_tab_tendencias(self) -> QWidget:
        widget = QWidget()
        layout = QVBoxLayout(widget)
        info = QLabel("Clique em uma data para abrir as notas daquele dia.")
        info.setStyleSheet("color: #64748b; font-size: 11px;")
        layout.addWidget(info)
        self.table_tendencias = QTableWidget()
        self.table_tendencias.setColumnCount(3)
        self.table_tendencias.setHorizontalHeaderLabels(["Data", "Valor Total", "Qtd Notas"])
        self._configurar_tabela(self.table_tendencias)
        self.table_tendencias.cellClicked.connect(self._abrir_detalhe_tendencia)
        layout.addWidget(self.table_tendencias)
        return widget

    def _criar_tab_cancelamentos(self) -> QWidget:
        widget = QWidget()
        layout = QVBoxLayout(widget)
        info = QLabel("Se houver canceladas, elas aparecem aqui. Clique em uma delas para ver os detalhes.")
        info.setStyleSheet("color: #64748b; font-size: 11px;")
        layout.addWidget(info)
        self.table_cancelamentos = QTableWidget()
        self.table_cancelamentos.setColumnCount(6)
        self.table_cancelamentos.setHorizontalHeaderLabels(["Número", "Data", "Cliente", "Serviço", "Valor", "Chave"])
        self._configurar_tabela(self.table_cancelamentos)
        self.table_cancelamentos.cellClicked.connect(self._abrir_detalhe_cancelamento)
        layout.addWidget(self.table_cancelamentos)
        return widget

    def _format_currency(self, valor) -> str:
        try:
            numero = float(valor or 0)
        except Exception:
            numero = 0.0
        return f"R$ {numero:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")

    def _format_data(self, data_valor) -> str:
        if not data_valor:
            return ""
        if isinstance(data_valor, datetime):
            return data_valor.strftime("%d/%m/%Y")
        if isinstance(data_valor, date):
            return data_valor.strftime("%d/%m/%Y")
        texto = str(data_valor).strip()
        for formato in ("%Y-%m-%d", "%Y-%m-%d %H:%M:%S", "%d/%m/%Y", "%d/%m/%Y %H:%M:%S"):
            try:
                return datetime.strptime(texto[:19], formato).strftime("%d/%m/%Y")
            except Exception:
                pass
        if 'T' in texto:
            try:
                return datetime.fromisoformat(texto.replace('Z', '+00:00')).strftime("%d/%m/%Y")
            except Exception:
                pass
        return texto

    def _valor_imposto(self, nota, *campos) -> float:
        for campo in campos:
            try:
                valor = float(getattr(nota, campo, 0) or 0)
            except Exception:
                valor = 0.0
            if valor > 0:
                return valor
        return 0.0

    def _only_digits(self, value) -> str:
        return ''.join(ch for ch in str(value or '') if ch.isdigit())

    def _papel_nota(self, nota) -> str:
        tipo_documento = str(getattr(nota, "tipo_documento", "") or "").upper()
        empresa_cnpj = self._only_digits(getattr(self, "empresa_cnpj", ""))
        empresa_nome = str(getattr(self, "empresa_nome", "") or "").strip().upper()
        emitente_cnpj = self._only_digits(getattr(nota, "emitente_cnpj", None))
        destinatario_cnpj = self._only_digits(getattr(nota, "destinatario_cnpj", None))
        emitente_nome = str(getattr(nota, "emitente_nome", None) or "").strip().upper()
        destinatario_nome = str(getattr(nota, "destinatario_nome", None) or "").strip().upper()

        if empresa_cnpj:
            if emitente_cnpj and emitente_cnpj == empresa_cnpj and destinatario_cnpj != empresa_cnpj:
                return "prestada"
            if destinatario_cnpj and destinatario_cnpj == empresa_cnpj and emitente_cnpj != empresa_cnpj:
                return "tomada"
            if emitente_cnpj == empresa_cnpj:
                return "prestada"
            if destinatario_cnpj == empresa_cnpj:
                return "tomada"

        if empresa_nome:
            if emitente_nome and emitente_nome == empresa_nome and destinatario_nome != empresa_nome:
                return "prestada"
            if destinatario_nome and destinatario_nome == empresa_nome and emitente_nome != empresa_nome:
                return "tomada"
            if emitente_nome == empresa_nome:
                return "prestada"
            if destinatario_nome == empresa_nome:
                return "tomada"

        if "PRESTAD" in tipo_documento or "SAÍDA" in tipo_documento or "SAIDA" in tipo_documento:
            return "prestada"
        if "TOMADA" in tipo_documento or "ENTRADA" in tipo_documento:
            return "tomada"
        return "desconhecido"

    def _texto_cliente(self, nota) -> str:
        if self._papel_nota(nota) == "tomada":
            return getattr(nota, "emitente_nome", None) or getattr(nota, "destinatario_nome", None) or "Desconhecido"
        return (
            getattr(nota, "destinatario_nome", None)
            or getattr(nota, "tomador_razao_social", None)
            or getattr(nota, "emitente_nome", None)
            or "Desconhecido"
        )

    def _texto_fornecedor(self, nota) -> str:
        return (
            getattr(nota, "emitente_nome", None)
            or getattr(nota, "destinatario_nome", None)
            or "Desconhecido"
        )

    def _texto_servico(self, nota) -> str:
        # Priorizar descrição real do serviço
        descricao = (
            getattr(nota, "descricao_servico", None)
            or getattr(nota, "servico_descricao", None)
        )
        if descricao and descricao.strip() and descricao not in ("NFS-e Prestada", "NFS-e Tomada"):
            return descricao
        # Se não houver descrição, usar tipo de documento
        tipo = getattr(nota, "tipo_documento", None)
        if tipo and tipo.strip():
            return tipo
        # Fallback para papel da nota
        papel = self._papel_nota(nota)
        if papel == "prestada":
            return "NFS-e Prestada"
        if papel == "tomada":
            return "NFS-e Tomada"
        return "Serviço Genérico"

    def _nota_cancelada(self, nota) -> bool:
        status = str(getattr(nota, "status", "") or getattr(nota, "situacao", "")).upper()
        return "CANCEL" in status or bool(getattr(nota, "status_cancelada", False))

    def _obter_tabela_aba_atual(self):
        aba_atual = self.tabs.currentWidget()
        mapa = {
            self.tab_financeiro: self.table_financeiro,
            self.tab_clientes: self.table_clientes,
            self.tab_fornecedores: self.table_fornecedores,
            self.tab_servicos: self.table_servicos,
            self.tab_impostos: self.table_impostos,
            self.tab_tendencias: self.table_tendencias,
            self.tab_cancelamentos: self.table_cancelamentos,
        }
        return mapa.get(aba_atual)

    def _aplicar_pesquisa_aba_atual(self, *_args):
        table = self._obter_tabela_aba_atual()
        if table is None:
            return

        termo = (self.search_input.text() or "").strip().lower()
        for row in range(table.rowCount()):
            mostrar = not termo
            if termo:
                for col in range(table.columnCount()):
                    item = table.item(row, col)
                    if item and termo in (item.text() or "").lower():
                        mostrar = True
                        break
            table.setRowHidden(row, not mostrar)

        if termo:
            self.status_label.setText(f"🔎 Pesquisa aplicada na aba atual: {self.search_input.text()}")
            self.status_label.setStyleSheet("color: #2563eb; font-weight: bold;")
        elif self.dados_notas:
            self.status_label.setText("✅ Relatórios gerados com sucesso!")
            self.status_label.setStyleSheet("color: #4CAF50; font-weight: bold;")
        else:
            self.status_label.setText("Pronto para gerar relatórios")
            self.status_label.setStyleSheet("color: #4CAF50; font-weight: bold;")

    def _limpar_pesquisa(self):
        self.search_input.clear()
        self._aplicar_pesquisa_aba_atual()

    def _filtrar_tabela(self, table: QTableWidget, termo: str):
        termo = (termo or "").strip().lower()
        for row in range(table.rowCount()):
            mostrar = not termo
            if termo:
                for col in range(table.columnCount()):
                    item = table.item(row, col)
                    if item and termo in (item.text() or "").lower():
                        mostrar = True
                        break
            table.setRowHidden(row, not mostrar)

    def _exportar_notas_dialogo_excel(self, titulo: str, notas: list):
        if not notas:
            QMessageBox.warning(self, "Validação", "Não há notas para exportar.")
            return

        if not OPENPYXL_AVAILABLE:
            QMessageBox.critical(self, "Erro", "openpyxl não está instalado")
            return

        nome_limpo = "".join(ch if ch.isalnum() or ch in (' ', '-', '_') else '_' for ch in titulo).strip().replace(' ', '_')
        caminho_padrao = Path.cwd() / f"{nome_limpo}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        caminho, _ = QFileDialog.getSaveFileName(
            self,
            "Salvar detalhes em Excel",
            str(caminho_padrao),
            "Excel Files (*.xlsx)"
        )
        if not caminho:
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "Notas"
        headers = [
            "Número", "Data", "Cliente", "Emitente", "Destinatário",
            "Serviço", "Valor", "Status", "Chave"
        ]
        for col, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col, value=header).font = Font(bold=True)

        for row, nota in enumerate(notas, start=2):
            valores = [
                str(getattr(nota, "numero", "") or ""),
                self._format_data(getattr(nota, "data_emissao", None)),
                self._texto_cliente(nota),
                str(getattr(nota, "emitente_nome", "") or ""),
                str(getattr(nota, "destinatario_nome", "") or ""),
                self._texto_servico(nota),
                self._format_currency(getattr(nota, "valor_total", 0)),
                str(getattr(nota, "status", "") or getattr(nota, "situacao", "") or ""),
                str(getattr(nota, "chave", "") or ""),
            ]
            for col, valor in enumerate(valores, start=1):
                ws.cell(row=row, column=col, value=valor)

        wb.save(caminho)
        QMessageBox.information(self, "Sucesso", f"Detalhes exportados para:\n{caminho}")

    def _mostrar_dialogo_notas(self, titulo: str, notas: list, subtitulo: str | None = None):
        dialog = QDialog(self)
        dialog.setWindowTitle(titulo)
        dialog.resize(1280, 620)

        layout = QVBoxLayout(dialog)
        if subtitulo:
            label = QLabel(subtitulo)
            label.setWordWrap(True)
            label.setStyleSheet("color: #64748b; font-size: 12px;")
            layout.addWidget(label)

        total = sum(float(getattr(nota, "valor_total", 0) or 0) for nota in notas)
        resumo = QLabel(f"Total de notas: {len(notas)} | Valor total: {self._format_currency(total)}")
        resumo.setStyleSheet("font-weight: 600; color: #0f172a;")
        layout.addWidget(resumo)

        info = QLabel("Dica: pesquise dentro desta janela e dê duplo clique em uma nota para abrir o XML/PDF do documento.")
        info.setStyleSheet("color: #64748b; font-size: 11px;")
        layout.addWidget(info)

        barra = QHBoxLayout()
        barra.addWidget(QLabel("Pesquisa nesta lista:"))
        search_dialog = QLineEdit()
        search_dialog.setPlaceholderText("Digite número, cliente, chave, serviço, valor, status...")
        barra.addWidget(search_dialog, 1)
        btn_limpar = QPushButton("Limpar")
        barra.addWidget(btn_limpar)
        btn_excel = QPushButton("📥 Exportar lista para Excel")
        barra.addWidget(btn_excel)
        layout.addLayout(barra)

        table = QTableWidget()
        table.setColumnCount(9)
        table.setHorizontalHeaderLabels([
            "Número", "Data", "Cliente", "Emitente", "Destinatário",
            "Serviço", "Valor", "Status", "Chave"
        ])
        self._configurar_tabela(table)
        table.setRowCount(len(notas))

        for idx, nota in enumerate(notas):
            valores = [
                str(getattr(nota, "numero", "") or ""),
                self._format_data(getattr(nota, "data_emissao", None)),
                self._texto_cliente(nota),
                str(getattr(nota, "emitente_nome", "") or ""),
                str(getattr(nota, "destinatario_nome", "") or ""),
                self._texto_servico(nota),
                self._format_currency(getattr(nota, "valor_total", 0)),
                str(getattr(nota, "status", "") or getattr(nota, "situacao", "") or ""),
                str(getattr(nota, "chave", "") or ""),
            ]
            for col, valor in enumerate(valores):
                table.setItem(idx, col, QTableWidgetItem(valor))

        def abrir_documento(row: int, _column: int):
            if 0 <= row < len(notas):
                self._abrir_arquivo_da_nota(notas[row])

        table.cellDoubleClicked.connect(abrir_documento)
        search_dialog.textChanged.connect(lambda texto: self._filtrar_tabela(table, texto))
        btn_limpar.clicked.connect(search_dialog.clear)
        btn_excel.clicked.connect(lambda: self._exportar_notas_dialogo_excel(titulo, notas))
        layout.addWidget(table)

        botoes = QDialogButtonBox(QDialogButtonBox.Close)
        botoes.rejected.connect(dialog.reject)
        botoes.accepted.connect(dialog.accept)
        botoes.button(QDialogButtonBox.Close).clicked.connect(dialog.accept)
        layout.addWidget(botoes)
        dialog.exec()

    def _abrir_arquivo_da_nota(self, nota):
        if not self.empresa_selecionada_id:
            QMessageBox.warning(self, "Aviso", "Nenhuma empresa selecionada.")
            return

        empresa = self.empresa_repo.get_by_id(self.empresa_selecionada_id)
        ok, destino = open_document_file(nota, empresa, self.cred_repo)
        if not ok:
            QMessageBox.warning(self, "Arquivo não localizado", destino)

    def _abrir_detalhe_cliente(self, row: int, _column: int):
        if row < 0 or row >= len(self._resultados_cliente):
            return
        cliente = self._resultados_cliente[row]
        nome = cliente["cliente"]
        notas = [
            nota for nota in self.dados_notas
            if self._papel_nota(nota) == "prestada" and self._texto_cliente(nota) == nome
        ]
        self._mostrar_dialogo_notas(
            f"Notas do cliente: {nome}",
            notas,
            "Lista das notas que compõem o total do cliente selecionado.",
        )


    def _abrir_detalhe_fornecedor(self, row: int, _column: int):
        if row < 0 or row >= len(self._resultados_fornecedor):
            return
        fornecedor = self._resultados_fornecedor[row]
        nome = fornecedor["fornecedor"]
        notas = [
            nota for nota in self.dados_notas
            if self._papel_nota(nota) == "tomada" and self._texto_fornecedor(nota) == nome
        ]
        self._mostrar_dialogo_notas(
            f"Notas do fornecedor: {nome}",
            notas,
            "Lista das notas tomadas que compõem o total do fornecedor selecionado.",
        )

    def _abrir_detalhe_servico(self, row: int, _column: int):
        if row < 0 or row >= len(self._resultados_servico):
            return
        servico = self._resultados_servico[row]
        nome = servico["servico"]
        notas = [nota for nota in self.dados_notas if self._texto_servico(nota) == nome]
        self._mostrar_dialogo_notas(
            f"Notas do serviço: {nome}",
            notas,
            "Lista completa das notas vinculadas ao serviço selecionado.",
        )

    def _abrir_detalhe_imposto(self, row: int, _column: int):
        if row < 0 or row >= len(self._resultados_impostos):
            return
        imposto = self._resultados_impostos[row]
        campos = tuple(imposto.get("campos") or (imposto["campo"],))
        notas = [nota for nota in self.dados_notas if self._valor_imposto(nota, *campos) > 0]
        self._mostrar_dialogo_notas(
            f"Notas com imposto: {imposto['imposto']}",
            notas,
            "Lista de notas que possuem valor lançado para o imposto selecionado.",
        )

    def _abrir_detalhe_tendencia(self, row: int, _column: int):
        if row < 0 or row >= len(self._resultados_tendencias):
            return
        tendencia = self._resultados_tendencias[row]
        data_referencia = tendencia["data"]
        notas = []
        for nota in self.dados_notas:
            data_nota = getattr(nota, 'data_emissao', None)
            if isinstance(data_nota, datetime):
                data_nota = data_nota.date()
            if data_nota == data_referencia:
                notas.append(nota)
        self._mostrar_dialogo_notas(
            f"Notas da data: {self._format_data(data_referencia)}",
            notas,
            "Lista de notas emitidas na data selecionada.",
        )

    def _abrir_detalhe_cancelamento(self, row: int, _column: int):
        if row < 0 or row >= len(self._resultados_cancelamentos):
            return
        nota = self._resultados_cancelamentos[row]
        identificador = getattr(nota, "numero", None) or getattr(nota, "chave", None) or "Documento cancelado"
        self._mostrar_dialogo_notas(
            f"Detalhes do cancelamento: {identificador}",
            [nota],
            "Documento cancelado encontrado dentro do período selecionado. Dê duplo clique para abrir o XML/PDF.",
        )

    def _on_global_empresa_changed(self, empresa_id):
        self.empresa_selecionada_id = empresa_id
        self._pending_auto_refresh = True
        if empresa_id:
            try:
                empresa = self.empresa_repo.get_by_id(empresa_id)
                if empresa:
                    self.empresa_label.setText(f"{empresa.razao_social} - {empresa.cnpj}")
                    self.empresa_cnpj = empresa.cnpj or ""
                    self.empresa_nome = empresa.razao_social or ""
                    self._save_state()
                    if self.isVisible() and not self._restoring_state and self._user_generated_once:
                        self._gerar_relatorios_silencioso()
                        self._pending_auto_refresh = False
                    elif not self._user_generated_once:
                        self.status_label.setText("Selecione o período e clique em 'Gerar Relatórios'.")
                        self.status_label.setStyleSheet("color: #64748b; font-weight: bold;")
                    return
            except Exception as exc:
                log.error(f"Erro ao atualizar empresa: {exc}")
        self.empresa_label.setText("Nenhuma empresa selecionada")
        self.empresa_cnpj = ""
        self.empresa_nome = ""

    def on_page_activated(self):
        if self._pending_auto_refresh and self._user_generated_once and self.empresa_selecionada_id:
            self._gerar_relatorios_silencioso()
            self._pending_auto_refresh = False

    def _save_state(self):
        if self.empresa_selecionada_id:
            self.settings.setValue("relatorios/empresa", str(self.empresa_selecionada_id))
        save_qdate(self.settings, "relatorios/data_inicial", self.data_inicial.date())
        save_qdate(self.settings, "relatorios/data_final", self.data_final.date())

    def _restore_state(self):
        self._restoring_state = True
        try:
            self.data_inicial.setDate(load_qdate(self.settings, "relatorios/data_inicial", QDate.currentDate().addMonths(-1)))
            self.data_final.setDate(load_qdate(self.settings, "relatorios/data_final", QDate.currentDate()))
            raw_empresa = self.settings.value("main/selected_company_id")
            empresa_id = None
            if raw_empresa not in (None, "", "None"):
                try:
                    empresa_id = int(raw_empresa)
                except Exception:
                    empresa_id = None
        finally:
            self._restoring_state = False
        if empresa_id:
            self._on_global_empresa_changed(empresa_id)

    def _as_report_item(self, documento):
        class _ReportItem(dict):
            def __getattr__(self, item):
                try:
                    return self[item]
                except KeyError as exc:
                    raise AttributeError(item) from exc

            __setattr__ = dict.__setitem__
            __delattr__ = dict.__delitem__

        tipo_documento = (getattr(documento, "tipo_documento", "") or "").strip()
        situacao = (getattr(documento, "situacao", "") or "").strip()
        status_base = (getattr(documento, "status", "") or "").strip() or situacao or "VALIDO"
        status_upper = status_base.upper()
        situacao_upper = situacao.upper()
        cancelada = "CANCEL" in status_upper or "CANCEL" in situacao_upper
        valor_total = float(getattr(documento, "valor_total", 0) or 0)

        cliente = getattr(documento, "destinatario_nome", None) or getattr(documento, "emitente_nome", None) or "Desconhecido"
        if "TOMADA" in tipo_documento.upper() or "ENTRADA" in tipo_documento.upper():
            cliente = getattr(documento, "emitente_nome", None) or getattr(documento, "destinatario_nome", None) or "Desconhecido"

        servico_descricao = tipo_documento or getattr(documento, "schema", None) or getattr(documento, "modelo", None) or "Documento fiscal"
        data_emissao = getattr(documento, "data_emissao", None)
        if isinstance(data_emissao, date) and not hasattr(data_emissao, 'hour'):
            data_emissao = datetime.combine(data_emissao, time.min)

        return _ReportItem(
            id=getattr(documento, "id", None),
            empresa_id=getattr(documento, "empresa_id", None),
            tipo_documento=tipo_documento,
            chave=getattr(documento, "chave", None),
            numero=getattr(documento, "numero", None),
            serie=getattr(documento, "serie", None),
            modelo=getattr(documento, "modelo", None),
            data_emissao=data_emissao,
            emitente_cnpj=getattr(documento, "emitente_cnpj", None),
            emitente_nome=getattr(documento, "emitente_nome", None),
            destinatario_cnpj=getattr(documento, "destinatario_cnpj", None),
            destinatario_nome=getattr(documento, "destinatario_nome", None),
            valor_total=valor_total,
            valor_servico=valor_total,
            situacao=situacao or status_base,
            status="CANCELADA" if cancelada else (situacao or status_base or "VALIDO"),
            status_cancelada=cancelada,
            tomador_razao_social=cliente,
            descricao_servico=servico_descricao,
            servico_descricao=servico_descricao,
            valor_issqn=0.0,
            valor_ir=0.0,
            valor_pis=0.0,
            valor_cofins=0.0,
            valor_csll=0.0,
            issqn_valor=0.0,
            irrf_valor=0.0,
            pis_valor=0.0,
            cofins_valor=0.0,
            csll_valor=0.0,
            origem_captura=getattr(documento, "origem_captura", None),
            schema=getattr(documento, "schema", None),
            arquivo_xml=getattr(documento, "arquivo_xml", None),
        )


    def _gerar_relatorios_silencioso(self):
        if not self._user_generated_once:
            return
        if not self.empresa_selecionada_id:
            return
        try:
            self.dados_notas = self._carregar_dados_relatorios(
                self.empresa_selecionada_id,
                self.data_inicial.date().toPython(),
                self.data_final.date().toPython(),
            )
            self._atualizar_tab_financeiro()
            self._atualizar_tab_clientes()
            self._atualizar_tab_fornecedores()
            self._atualizar_tab_servicos()
            self._atualizar_tab_impostos()
            self._atualizar_tab_tendencias()
            self._atualizar_tab_cancelamentos()
            self.btn_exportar.setEnabled(bool(self.dados_notas))
            if self.search_input.text().strip():
                self._aplicar_pesquisa_aba_atual()
            elif self.dados_notas:
                self.status_label.setText("✅ Relatórios atualizados automaticamente!")
                self.status_label.setStyleSheet("color: #4CAF50; font-weight: bold;")
        except Exception as exc:
            log.error(f"Falha ao atualizar relatórios automaticamente: {exc}")

    def _carregar_dados_relatorios(self, empresa_id, data_ini, data_fim):
        metodo_periodo = getattr(self.doc_repo, "get_by_empresa_and_period", None)
        if callable(metodo_periodo):
            return metodo_periodo(empresa_id, data_ini, data_fim)

        log.warning(
            "DocumentoRepository sem get_by_empresa_and_period; usando fallback compatível na própria tela de relatórios."
        )
        documentos = self.doc_repo.list_by_empresa(empresa_id)
        itens = []
        for documento in documentos:
            data_emissao = getattr(documento, "data_emissao", None)
            if isinstance(data_emissao, datetime):
                data_base = data_emissao.date()
            else:
                data_base = data_emissao
            if data_ini and data_base and data_base < data_ini:
                continue
            if data_fim and data_base and data_base > data_fim:
                continue
            itens.append(self._as_report_item(documento))
        itens.sort(key=lambda item: ((item.data_emissao or datetime.min), item.id or 0))
        return itens

    def _gerar_relatorios(self):
        self._user_generated_once = True
        if not self.empresa_selecionada_id:
            QMessageBox.warning(self, "Validação", "Selecione uma empresa no topo da tela")
            return

        self.progress.setVisible(True)
        self.progress.setValue(0)
        self.btn_gerar.setEnabled(False)
        self.status_label.setText("Gerando relatórios...")
        self.status_label.setStyleSheet("color: #2563eb;")

        try:
            empresa_id = self.empresa_selecionada_id
            data_ini = self.data_inicial.date().toPython()
            data_fim = self.data_final.date().toPython()

            self.dados_notas = self._carregar_dados_relatorios(empresa_id, data_ini, data_fim)
            self.progress.setValue(20)

            self._atualizar_tab_financeiro()
            self.progress.setValue(35)

            self._atualizar_tab_clientes()
            self.progress.setValue(50)

            self._atualizar_tab_fornecedores()
            self.progress.setValue(60)

            self._atualizar_tab_servicos()
            self.progress.setValue(70)

            self._atualizar_tab_impostos()
            self.progress.setValue(80)

            self._atualizar_tab_tendencias()
            self.progress.setValue(90)

            self._atualizar_tab_cancelamentos()
            self.progress.setValue(100)

            self.btn_exportar.setEnabled(True)
            self.status_label.setText("✅ Relatórios gerados com sucesso!")
            self.status_label.setStyleSheet("color: #4CAF50;")
            self._aplicar_pesquisa_aba_atual()

        except Exception as exc:
            log.exception(f"Erro ao gerar relatórios: {exc}")
            self.status_label.setText(f"❌ Erro: {exc}")
            self.status_label.setStyleSheet("color: #b91c1c;")
            QMessageBox.critical(self, "Erro", f"Erro ao gerar relatórios: {exc}")

        finally:
            self.progress.setVisible(False)
            self.btn_gerar.setEnabled(True)

    def _atualizar_tab_financeiro(self):
        if not self.dados_notas:
            self.table_financeiro.setRowCount(0)
            return

        total_notas = len(self.dados_notas)
        total_valor = sum(float(n.valor_servico or 0) for n in self.dados_notas)
        valor_medio = total_valor / total_notas if total_notas > 0 else 0
        valor_minimo = min((float(n.valor_servico or 0) for n in self.dados_notas), default=0)
        valor_maximo = max((float(n.valor_servico or 0) for n in self.dados_notas), default=0)
        canceladas = sum(1 for n in self.dados_notas if n.status_cancelada)
        taxa_cancelamento = (canceladas / total_notas * 100) if total_notas > 0 else 0

        self.table_financeiro.setRowCount(6)
        rows = [
            ("Total de Notas", str(total_notas)),
            ("Valor Total", self._format_currency(total_valor)),
            ("Valor Médio", self._format_currency(valor_medio)),
            ("Valor Mínimo", self._format_currency(valor_minimo)),
            ("Valor Máximo", self._format_currency(valor_maximo)),
            ("Taxa de Cancelamento", f"{taxa_cancelamento:.1f}%"),
        ]
        for idx, (metrica, valor) in enumerate(rows):
            self.table_financeiro.setItem(idx, 0, QTableWidgetItem(metrica))
            self.table_financeiro.setItem(idx, 1, QTableWidgetItem(valor))

    def _atualizar_tab_clientes(self):
        if not self.dados_notas:
            self._resultados_cliente = []
            self.table_clientes.setRowCount(0)
            return

        clientes = defaultdict(lambda: {"valor": 0.0, "qtd": 0})
        for nota in self.dados_notas:
            if self._papel_nota(nota) != "prestada":
                continue
            cliente = self._texto_cliente(nota)
            clientes[cliente]["valor"] += float(nota.valor_servico or 0)
            clientes[cliente]["qtd"] += 1

        self._resultados_cliente = [
            {
                "cliente": cliente,
                "valor_total": dados["valor"],
                "quantidade_notas": dados["qtd"],
                "valor_medio": (dados["valor"] / dados["qtd"]) if dados["qtd"] else 0,
            }
            for cliente, dados in sorted(clientes.items(), key=lambda item: item[1]["valor"], reverse=True)
        ]

        self.table_clientes.setRowCount(len(self._resultados_cliente))
        for idx, cliente in enumerate(self._resultados_cliente):
            self.table_clientes.setItem(idx, 0, QTableWidgetItem(cliente["cliente"]))
            self.table_clientes.setItem(idx, 1, QTableWidgetItem(self._format_currency(cliente["valor_total"])))
            self.table_clientes.setItem(idx, 2, QTableWidgetItem(str(cliente["quantidade_notas"])))
            self.table_clientes.setItem(idx, 3, QTableWidgetItem(self._format_currency(cliente["valor_medio"])))


    def _atualizar_tab_fornecedores(self):
        if not self.dados_notas:
            self._resultados_fornecedor = []
            self.table_fornecedores.setRowCount(0)
            return

        fornecedores = defaultdict(lambda: {"valor": 0.0, "qtd": 0})
        for nota in self.dados_notas:
            if self._papel_nota(nota) != "tomada":
                continue
            fornecedor = self._texto_fornecedor(nota)
            fornecedores[fornecedor]["valor"] += float(nota.valor_servico or 0)
            fornecedores[fornecedor]["qtd"] += 1

        self._resultados_fornecedor = [
            {
                "fornecedor": fornecedor,
                "valor_total": dados["valor"],
                "quantidade_notas": dados["qtd"],
                "valor_medio": (dados["valor"] / dados["qtd"]) if dados["qtd"] else 0,
            }
            for fornecedor, dados in sorted(fornecedores.items(), key=lambda item: item[1]["valor"], reverse=True)
        ]

        self.table_fornecedores.setRowCount(len(self._resultados_fornecedor))
        for idx, fornecedor in enumerate(self._resultados_fornecedor):
            self.table_fornecedores.setItem(idx, 0, QTableWidgetItem(fornecedor["fornecedor"]))
            self.table_fornecedores.setItem(idx, 1, QTableWidgetItem(self._format_currency(fornecedor["valor_total"])))
            self.table_fornecedores.setItem(idx, 2, QTableWidgetItem(str(fornecedor["quantidade_notas"])))
            self.table_fornecedores.setItem(idx, 3, QTableWidgetItem(self._format_currency(fornecedor["valor_medio"])))

    def _atualizar_tab_servicos(self):
        if not self.dados_notas:
            self._resultados_servico = []
            self.table_servicos.setRowCount(0)
            return

        servicos = defaultdict(lambda: {"valor": 0.0, "qtd": 0})
        for nota in self.dados_notas:
            servico = self._texto_servico(nota)
            servicos[servico]["valor"] += float(nota.valor_servico or 0)
            servicos[servico]["qtd"] += 1

        self._resultados_servico = [
            {
                "servico": servico,
                "valor_total": dados["valor"],
                "quantidade_notas": dados["qtd"],
            }
            for servico, dados in sorted(servicos.items(), key=lambda item: item[1]["valor"], reverse=True)
        ]

        self.table_servicos.setRowCount(len(self._resultados_servico))
        for idx, servico in enumerate(self._resultados_servico):
            self.table_servicos.setItem(idx, 0, QTableWidgetItem(servico["servico"]))
            self.table_servicos.setItem(idx, 1, QTableWidgetItem(self._format_currency(servico["valor_total"])))
            self.table_servicos.setItem(idx, 2, QTableWidgetItem(str(servico["quantidade_notas"])))

    def _atualizar_tab_impostos(self):
        if not self.dados_notas:
            self._resultados_impostos = []
            self.table_impostos.setRowCount(0)
            return

        configuracoes = [
            ("ISSQN", ("valor_issqn", "issqn_valor")),
            ("IRRF", ("valor_ir", "valor_irrf", "irrf_valor")),
            ("PIS", ("valor_pis", "pis_valor")),
            ("COFINS", ("valor_cofins", "cofins_valor")),
            ("CSLL", ("valor_csll", "csll_valor")),
        ]
        self._resultados_impostos = []
        for nome, campos in configuracoes:
            notas = [nota for nota in self.dados_notas if self._valor_imposto(nota, *campos) > 0]
            valor_total = sum(self._valor_imposto(nota, *campos) for nota in notas)
            self._resultados_impostos.append({
                "imposto": nome,
                "campo": campos[0],
                "campos": campos,
                "valor_total": valor_total,
                "quantidade_notas": len(notas),
            })

        self.table_impostos.setRowCount(len(self._resultados_impostos))
        for idx, registro in enumerate(self._resultados_impostos):
            self.table_impostos.setItem(idx, 0, QTableWidgetItem(registro["imposto"]))
            self.table_impostos.setItem(idx, 1, QTableWidgetItem(self._format_currency(registro["valor_total"])))
            self.table_impostos.setItem(idx, 2, QTableWidgetItem(str(registro["quantidade_notas"])))

    def _atualizar_tab_tendencias(self):
        if not self.dados_notas:
            self._resultados_tendencias = []
            self.table_tendencias.setRowCount(0)
            return

        agrupado = defaultdict(lambda: {"valor": 0.0, "qtd": 0})
        for nota in self.dados_notas:
            data_nota = getattr(nota, 'data_emissao', None)
            if isinstance(data_nota, datetime):
                data_nota = data_nota.date()
            if not data_nota:
                continue
            agrupado[data_nota]["valor"] += float(nota.valor_servico or 0)
            agrupado[data_nota]["qtd"] += 1

        self._resultados_tendencias = [
            {"data": data_ref, "valor_total": dados["valor"], "quantidade_notas": dados["qtd"]}
            for data_ref, dados in sorted(agrupado.items())
        ]

        self.table_tendencias.setRowCount(len(self._resultados_tendencias))
        for idx, registro in enumerate(self._resultados_tendencias):
            self.table_tendencias.setItem(idx, 0, QTableWidgetItem(self._format_data(registro["data"])))
            self.table_tendencias.setItem(idx, 1, QTableWidgetItem(self._format_currency(registro["valor_total"])))
            self.table_tendencias.setItem(idx, 2, QTableWidgetItem(str(registro["quantidade_notas"])))

    def _atualizar_tab_cancelamentos(self):
        if not self.dados_notas:
            self._resultados_cancelamentos = []
            self.table_cancelamentos.setRowCount(0)
            return

        # Mostrar todas as notas canceladas, independentemente se são prestadas ou tomadas
        self._resultados_cancelamentos = [
            nota for nota in self.dados_notas
            if self._nota_cancelada(nota)
        ]
        self.table_cancelamentos.setRowCount(len(self._resultados_cancelamentos))
        for idx, nota in enumerate(self._resultados_cancelamentos):
            self.table_cancelamentos.setItem(idx, 0, QTableWidgetItem(str(nota.numero or "")))
            self.table_cancelamentos.setItem(idx, 1, QTableWidgetItem(self._format_data(nota.data_emissao)))
            self.table_cancelamentos.setItem(idx, 2, QTableWidgetItem(self._texto_cliente(nota)))
            self.table_cancelamentos.setItem(idx, 3, QTableWidgetItem(self._texto_servico(nota)))
            self.table_cancelamentos.setItem(idx, 4, QTableWidgetItem(self._format_currency(nota.valor_total)))
            self.table_cancelamentos.setItem(idx, 5, QTableWidgetItem(str(nota.chave or "")))

    def _exportar_excel(self):
        if not self.dados_notas:
            QMessageBox.warning(self, "Validação", "Gere os relatórios primeiro")
            return

        if not OPENPYXL_AVAILABLE:
            QMessageBox.critical(self, "Erro", "openpyxl não está instalado")
            return

        try:
            empresa_nome = self.empresa_label.text().split(" - ")[0] if " - " in self.empresa_label.text() else "Relatorio"
            caminho_padrao = Path.cwd() / f"Relatorio_{empresa_nome}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

            caminho, _ = QFileDialog.getSaveFileName(
                self,
                "Salvar Relatório",
                str(caminho_padrao),
                "Excel Files (*.xlsx)"
            )

            if not caminho:
                return

            wb = Workbook()
            wb.remove(wb.active)

            # Financeiro
            ws = wb.create_sheet("Financeiro")
            ws['A1'] = "RELATÓRIO FINANCEIRO"
            ws['A1'].font = Font(bold=True, size=12)
            for idx in range(self.table_financeiro.rowCount()):
                item_a = self.table_financeiro.item(idx, 0)
                item_b = self.table_financeiro.item(idx, 1)
                ws[f'A{idx + 3}'] = item_a.text() if item_a else ""
                ws[f'B{idx + 3}'] = item_b.text() if item_b else ""

            # Clientes
            ws = wb.create_sheet("Clientes")
            ws['A1'] = "RELATÓRIO DE CLIENTES"
            ws['A1'].font = Font(bold=True, size=12)
            headers = ["Cliente", "Valor Total", "Qtd Notas", "Valor Médio"]
            for c, header in enumerate(headers, start=1):
                ws.cell(row=3, column=c, value=header).font = Font(bold=True)
            for r, registro in enumerate(self._resultados_cliente, start=4):
                ws.cell(row=r, column=1, value=registro["cliente"])
                ws.cell(row=r, column=2, value=self._format_currency(registro["valor_total"]))
                ws.cell(row=r, column=3, value=registro["quantidade_notas"])
                ws.cell(row=r, column=4, value=self._format_currency(registro["valor_medio"]))

            # Fornecedores
            ws = wb.create_sheet("Fornecedores")
            ws['A1'] = "RELATÓRIO DE FORNECEDORES"
            ws['A1'].font = Font(bold=True, size=12)
            headers = ["Fornecedor", "Valor Total", "Qtd Notas", "Valor Médio"]
            for c, header in enumerate(headers, start=1):
                ws.cell(row=3, column=c, value=header).font = Font(bold=True)
            for r, registro in enumerate(self._resultados_fornecedor, start=4):
                ws.cell(row=r, column=1, value=registro["fornecedor"])
                ws.cell(row=r, column=2, value=self._format_currency(registro["valor_total"]))
                ws.cell(row=r, column=3, value=registro["quantidade_notas"])
                ws.cell(row=r, column=4, value=self._format_currency(registro["valor_medio"]))

            # Serviços
            ws = wb.create_sheet("Serviços")
            ws['A1'] = "RELATÓRIO DE SERVIÇOS"
            ws['A1'].font = Font(bold=True, size=12)
            headers = ["Serviço", "Valor Total", "Qtd Notas"]
            for c, header in enumerate(headers, start=1):
                ws.cell(row=3, column=c, value=header).font = Font(bold=True)
            for r, registro in enumerate(self._resultados_servico, start=4):
                ws.cell(row=r, column=1, value=registro["servico"])
                ws.cell(row=r, column=2, value=self._format_currency(registro["valor_total"]))
                ws.cell(row=r, column=3, value=registro["quantidade_notas"])

            # Impostos
            ws = wb.create_sheet("Impostos")
            ws['A1'] = "RELATÓRIO DE IMPOSTOS"
            ws['A1'].font = Font(bold=True, size=12)
            headers = ["Imposto", "Valor Total", "Qtd Notas"]
            for c, header in enumerate(headers, start=1):
                ws.cell(row=3, column=c, value=header).font = Font(bold=True)
            for r, registro in enumerate(self._resultados_impostos, start=4):
                ws.cell(row=r, column=1, value=registro["imposto"])
                ws.cell(row=r, column=2, value=self._format_currency(registro["valor_total"]))
                ws.cell(row=r, column=3, value=registro["quantidade_notas"])

            # Tendências
            ws = wb.create_sheet("Tendências")
            ws['A1'] = "RELATÓRIO DE TENDÊNCIAS"
            ws['A1'].font = Font(bold=True, size=12)
            headers = ["Data", "Valor Total", "Qtd Notas"]
            for c, header in enumerate(headers, start=1):
                ws.cell(row=3, column=c, value=header).font = Font(bold=True)
            for r, registro in enumerate(self._resultados_tendencias, start=4):
                ws.cell(row=r, column=1, value=self._format_data(registro["data"]))
                ws.cell(row=r, column=2, value=self._format_currency(registro["valor_total"]))
                ws.cell(row=r, column=3, value=registro["quantidade_notas"])

            # Cancelamentos
            ws = wb.create_sheet("Cancelamentos")
            ws['A1'] = "NOTAS CANCELADAS"
            ws['A1'].font = Font(bold=True, size=12)
            headers = ["Número", "Data", "Cliente", "Serviço", "Valor", "Chave"]
            for c, header in enumerate(headers, start=1):
                ws.cell(row=3, column=c, value=header).font = Font(bold=True)
            for r, nota in enumerate(self._resultados_cancelamentos, start=4):
                ws.cell(row=r, column=1, value=str(nota.numero or ""))
                ws.cell(row=r, column=2, value=self._format_data(nota.data_emissao))
                ws.cell(row=r, column=3, value=self._texto_cliente(nota))
                ws.cell(row=r, column=4, value=self._texto_servico(nota))
                ws.cell(row=r, column=5, value=self._format_currency(nota.valor_total))
                ws.cell(row=r, column=6, value=str(nota.chave or ""))

            wb.save(caminho)
            QMessageBox.information(self, "Sucesso", f"Relatório salvo em:\n{caminho}")
            log.info(f"Relatório exportado: {caminho}")

        except Exception as exc:
            log.exception(f"Erro ao exportar: {exc}")
            QMessageBox.critical(self, "Erro", f"Erro ao exportar: {exc}")
