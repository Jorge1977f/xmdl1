"""Exportador de relatórios para Excel com formatação profissional."""
from __future__ import annotations

from datetime import datetime
from pathlib import Path
from decimal import Decimal

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from app.utils.logger import log
from config.settings import DOWNLOADS_DIR


class ExcelExporter:
    """Exporta dados para Excel com formatação profissional."""

    def __init__(self, titulo: str = "Relatório"):
        """Inicializa o exportador."""
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl não está instalado. Execute: pip install openpyxl")

        self.titulo = titulo
        self.workbook = Workbook()
        self.workbook.remove(self.workbook.active)
        self.abas = {}

    def adicionar_aba(self, nome: str) -> dict:
        """Adiciona uma nova aba ao workbook."""
        ws = self.workbook.create_sheet(nome)
        self.abas[nome] = ws
        return ws

    def adicionar_relatorio_financeiro(self, relatorio) -> None:
        """Adiciona relatório financeiro."""
        ws = self.adicionar_aba("Financeiro")

        # Cabeçalho
        ws['A1'] = "RELATÓRIO FINANCEIRO"
        ws['A1'].font = Font(size=14, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        ws.merge_cells('A1:B1')

        # Dados
        dados = [
            ("Total de Notas", relatorio.total_notas),
            ("Valor Total", f"R$ {relatorio.total_valor:,.2f}"),
            ("Valor Médio", f"R$ {relatorio.valor_medio:,.2f}"),
            ("Valor Mínimo", f"R$ {relatorio.valor_minimo:,.2f}"),
            ("Valor Máximo", f"R$ {relatorio.valor_maximo:,.2f}"),
            ("Notas Válidas", relatorio.notas_validas),
            ("Notas Canceladas", relatorio.notas_canceladas),
            ("Taxa de Cancelamento", f"{relatorio.percentual_canceladas:.2f}%"),
        ]

        for idx, (label, valor) in enumerate(dados, start=3):
            ws[f'A{idx}'] = label
            ws[f'B{idx}'] = valor
            ws[f'A{idx}'].font = Font(bold=True)

        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20

    def adicionar_relatorio_clientes(self, relatorio) -> None:
        """Adiciona relatório de clientes."""
        ws = self.adicionar_aba("Clientes")

        # Cabeçalho
        ws['A1'] = "TOP 10 CLIENTES"
        ws['A1'].font = Font(size=14, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        ws.merge_cells('A1:D1')

        # Headers
        headers = ["Cliente", "Valor Total", "Qtd Notas", "Valor Médio"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

        # Dados
        for row_idx, cliente in enumerate(relatorio.top_clientes, start=4):
            ws.cell(row=row_idx, column=1).value = cliente.get('cliente', '')
            ws.cell(row=row_idx, column=2).value = f"R$ {cliente.get('valor_total', 0):,.2f}"
            ws.cell(row=row_idx, column=3).value = cliente.get('quantidade_notas', 0)
            ws.cell(row=row_idx, column=4).value = f"R$ {cliente.get('valor_medio', 0):,.2f}"

        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 15

    def adicionar_relatorio_servicos(self, relatorio) -> None:
        """Adiciona relatório de serviços."""
        ws = self.adicionar_aba("Serviços")

        # Cabeçalho
        ws['A1'] = "TOP SERVIÇOS"
        ws['A1'].font = Font(size=14, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        ws.merge_cells('A1:C1')

        # Headers
        headers = ["Serviço", "Valor Total", "Qtd Notas"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

        # Dados
        for row_idx, servico in enumerate(relatorio.top_servicos, start=4):
            ws.cell(row=row_idx, column=1).value = servico.get('servico', '')
            ws.cell(row=row_idx, column=2).value = f"R$ {servico.get('valor_total', 0):,.2f}"
            ws.cell(row=row_idx, column=3).value = servico.get('quantidade_notas', 0)

        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 12

    def adicionar_relatorio_impostos(self, relatorio) -> None:
        """Adiciona relatório de impostos."""
        ws = self.adicionar_aba("Impostos")

        # Cabeçalho
        ws['A1'] = "ANÁLISE DE IMPOSTOS"
        ws['A1'].font = Font(size=14, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        ws.merge_cells('A1:B1')

        # Dados
        dados = [
            ("Total ISSQN", f"R$ {relatorio.total_issqn:,.2f}"),
            ("Total IRRF", f"R$ {relatorio.total_irrf:,.2f}"),
            ("Total PIS", f"R$ {relatorio.total_pis:,.2f}"),
            ("Total COFINS", f"R$ {relatorio.total_cofins:,.2f}"),
            ("Total CSLL", f"R$ {relatorio.total_csll:,.2f}"),
            ("Total Retenções", f"R$ {relatorio.total_retencoes:,.2f}"),
            ("% ISSQN", f"{relatorio.percentual_issqn:.2f}%"),
            ("% Retenções", f"{relatorio.percentual_retencoes:.2f}%"),
        ]

        for idx, (label, valor) in enumerate(dados, start=3):
            ws[f'A{idx}'] = label
            ws[f'B{idx}'] = valor
            ws[f'A{idx}'].font = Font(bold=True)

        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20

    def adicionar_relatorio_tendencias(self, relatorio) -> None:
        """Adiciona relatório de tendências."""
        ws = self.adicionar_aba("Tendências")

        # Cabeçalho
        ws['A1'] = "ANÁLISE DE TENDÊNCIAS"
        ws['A1'].font = Font(size=14, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        ws.merge_cells('A1:B1')

        # Headers
        ws['A3'] = "Data"
        ws['B3'] = "Valor"
        for cell in [ws['A3'], ws['B3']]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

        # Dados
        for row_idx, periodo in enumerate(relatorio.periodos, start=4):
            ws.cell(row=row_idx, column=1).value = periodo.get('data', '')
            ws.cell(row=row_idx, column=2).value = f"R$ {periodo.get('valor', 0):,.2f}"

        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 15

    def adicionar_relatorio_cancelamentos(self, relatorio) -> None:
        """Adiciona relatório de cancelamentos."""
        ws = self.adicionar_aba("Cancelamentos")

        # Cabeçalho
        ws['A1'] = "ANÁLISE DE CANCELAMENTOS"
        ws['A1'].font = Font(size=14, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        ws.merge_cells('A1:B1')

        # Dados
        dados = [
            ("Total de Notas", relatorio.total_canceladas + relatorio.total_validas),
            ("Notas Válidas", relatorio.total_validas),
            ("Notas Canceladas", relatorio.total_canceladas),
            ("Taxa de Cancelamento", f"{relatorio.taxa_cancelamento:.2f}%"),
        ]

        for idx, (label, valor) in enumerate(dados, start=3):
            ws[f'A{idx}'] = label
            ws[f'B{idx}'] = valor
            ws[f'A{idx}'].font = Font(bold=True)

        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20

    def salvar(self, caminho: str | Path) -> bool:
        """Salva o workbook."""
        try:
            caminho = Path(caminho)
            caminho.parent.mkdir(parents=True, exist_ok=True)
            self.workbook.save(str(caminho))
            log.info(f"Relatório salvo: {caminho}")
            return True
        except Exception as exc:
            log.error(f"Erro ao salvar relatório: {exc}")
            return False

    def obter_caminho_padrao(self, empresa_nome: str = "Empresa") -> str:
        """Retorna caminho padrão para salvar."""
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        return str(Path(DOWNLOADS_DIR) / f"Relatorio_{empresa_nome}_{timestamp}.xlsx")
